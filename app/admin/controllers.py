from flask import request, jsonify, flash
from flask_login import logout_user
from sqlalchemy import inspect, or_, text
from sqlalchemy.exc import SQLAlchemyError
from app import db
from app.admin.email.engine import email_builder_factory, EmailSender
from app.models.customer import Customer
from app.models.mailing import Mailing
from app.models.notifications import Notifications
from app.models.job import Jobs
from app.models.quote import Quote
from .models import QuoteForm, CustomerForm, JobForm, PromoForm
import pandas as pd
import os
from datetime import datetime
import requests
from io import BytesIO
from typing import Any, List, Tuple, Dict
from http import HTTPStatus
from openpyxl import load_workbook
from pytz import utc
from sib_api_v3_sdk import Configuration, ApiClient
from sib_api_v3_sdk.api.account_api import AccountApi

def create_table() -> Tuple[Dict[str, Any], int]:
    """
    Creates database tables if they do not exist.

    :return: A tuple containing a dictionary with a message and table existence information, and an HTTP status code.
    """
    table_names = ['customers', 'mailing', 'notifications', 'jobs', 'quotes']
    existing_tables = inspect(db.engine).get_table_names()
    table_existence = {table: table in existing_tables for table in table_names}
    try:
        db.create_all()
        return {'message': 'Tables created or already exist', 'tables': table_existence}, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_sales_scoreboard(scoreboard_users: List[Dict[str, str]]):
    """
    Retrieves the sales scoreboard for the given users.

    :param scoreboard_users: List of users.
    :return: Sorted list of user scores.
    """
    the_score = []
    the_7_day_score = []
    seven_days_ago = (datetime.now() - pd.Timedelta(days=7))
    try:
        for user in scoreboard_users:
            score = len(Customer.query.filter_by(salesman=user['username']).all())
            obj = {"username": user['username'], "score": score, "call_sign": user['call_sign']}
            the_score.append(obj)

            seven_days_ago_str = seven_days_ago.strftime('%Y-%m-%d') 
            score_7 = len(Customer.query.filter_by(salesman=user['username']).filter(Customer.date >= seven_days_ago_str).all())
            obj_7 = {"username": user['username'], "score": score_7, "call_sign": user['call_sign']}
            the_7_day_score.append(obj_7)
        the_score = sorted(the_score, key=lambda k: k['score'], reverse=True)
        the_7_day_score = sorted(the_7_day_score, key=lambda k: k['score'], reverse=True)
    except SQLAlchemyError as e:
        flash("The tables do not exist. Please create them first.")
    return [the_score, the_7_day_score]

def delete_table(headers: Dict[str, str]) -> Tuple[Dict[str, str], int]:
    """
    Deletes all database tables. Used in development mode.

    :param headers: Request headers containing Authorization.
    :return: Message and HTTP status code.
    """
    Authorization = headers.get('Authorization')
    if Authorization == "Basic getalife:yesyou": # Consider using an environment variable here
        try:
            db.drop_all()
            return {'message': 'Table deleted'}, HTTPStatus.OK
        except SQLAlchemyError as e:
            return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR
    else:
        return {'message': 'An error occurred', 'error': 'Invalid credentials'}, HTTPStatus.INTERNAL_SERVER_ERROR

def check_service_status(url: str) -> bool:
    """
    Checks the status of a given service.

    :param url: URL of the service.
    :return: True if online, False otherwise.
    """
    try:
        response = requests.get(url)
        return response.status_code == HTTPStatus.OK
    except requests.RequestException:
        return False

def build_service_list(service_list: List[Dict[str, str]]) -> str:
    """
    Builds an HTML list of services and their statuses.

    :param service_list: List of services.
    :return: HTML string representing the service list.
    """
    services_ = ""
    for service in service_list:
        for service_name, url in service.items():
            status = "ONLINE" if check_service_status(url) else "OFFLINE"
            badge = "success" if status == "ONLINE" else "danger"
            services_ += f'<li><span class="badge badge-{badge}">{status}</span> : <a href="{url}">{service_name.capitalize()}</a></li>'
    return services_

def insert_into_promo_list(form) -> Tuple[Dict[str, str], int]:
    """
    Inserts a new record into the promotional mailing list.

    :param form: Form data containing customer and email information.
    :return: A message and HTTP status code.
    """
    try:
        if 'mailing' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR
        new_customer = Mailing(customer=form.customer.data, email=form.email.data)
        db.session.add(new_customer)
        db.session.commit()
        return {'message': 'Customer added'}, HTTPStatus.OK
    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def add_customer(form, welcome_email: bool = False, mailList: bool = False) -> Tuple[Dict[str, str], int]:
    """
    Adds a new customer and optionally adds them to the mailing list and sends a welcome email.

    :param form: Form data containing customer details.
    :param welcome_email: Boolean indicating whether to send a welcome email.
    :param mailList: Boolean indicating whether to add to the mailing list.
    :return: A message and HTTP status code.
    """
    if 'customers' not in inspect(db.engine).get_table_names():
        return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR
    try:
        if mailList:
            mail_promo_customer = Mailing(customer=form.name.data, email=form.email.data)
            db.session.add(mail_promo_customer)
        new_customer = Customer(contract_number=form.contract_number.data, salesman=form.salesman.data,
                                name=form.name.data, address=form.address.data, city=form.city.data,
                                phone=form.phone.data, description=form.description.data, date=form.date.data, email=form.email.data)
        db.session.add(new_customer)
        db.session.commit()
        conf = None
        if welcome_email:
            builder = email_builder_factory('welcome')
            email_sender = EmailSender(builder)
            conf = email_sender.send_email(new_customer.email, "welcome", new_customer.salesman)
        return {'message': 'Customer added', 'status': str(conf)}, HTTPStatus.OK
    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def delete_customer(id: int) -> Tuple[Dict[str, str], int]:
    """
    Deletes a customer by ID.

    :param id: Customer ID.
    :return: A message and HTTP status code.
    """
    try:
        rows_deleted = Customer.query.filter_by(id=id).delete()
        if rows_deleted == 0:
            return {'message': 'No customer found with this id'}, HTTPStatus.NOT_FOUND

        db.session.commit()
        return {'message': 'deleted'}, HTTPStatus.OK

    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_customers() -> Tuple[List[Customer], int]:
    """
    Retrieves all customers.

    :return: List of customers and HTTP status code.
    """
    try:
        if 'customers' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR

        customers = Customer.query.all()
        return customers, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_customers_based_on_params(params: Dict[str, str]) -> Tuple[List[Customer], int]:
    """
    Retrieves customers based on the given parameters.

    :param params: Filter parameters.
    :return: List of customers and HTTP status code.
    """
    try:
        if 'customers' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR

        query = Customer.query
        for key, value in params.items():
            if value.strip():
                if key == 'description':
                    query = query.filter(Customer.description.like(f'%{value}%'))
                else:
                    query = query.filter_by(**{key: value})

        customers = query.all()
        return customers, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_mail_list() -> Tuple[List[Mailing], int]:
    """
    Retrieve all customers from the mailing list.

    :return: List of customers in the mailing list and HTTP status code.
    """
    try:
        if 'mailing' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR
        
        customers = Mailing.query.all()
        return customers, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def send_promo_email(salesman: str, data: Dict[str, str]) -> None:
    """
    Send promotional emails to all customers in the mailing list.

    :param salesman: The salesman's name.
    :param data: Email data containing body and subject.
    """
    try:
        if 'mailing' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR

        customers = Mailing.query.all()
        for customer in customers:
            builder = email_builder_factory('mailing')
            email_sender = EmailSender(builder)
            conf = email_sender.send_email(customer.email, "mailing", salesman, data['emailBody'], data['emailSubject'])
    except SQLAlchemyError:
        pass

def delete_mail_promo(id: int) -> Tuple[Dict[str, str], int]:
    """
    Deletes a customer from the mailing promo list by ID.

    :param id: Customer ID in the mailing list.
    :return: A message and HTTP status code.
    """
    try:
        rows_deleted = Mailing.query.filter_by(id=id).delete()
        if rows_deleted == 0:
            return {'message': 'No customer found with this id'}, HTTPStatus.NOT_FOUND

        db.session.commit()
        return {'message': 'deleted'}, HTTPStatus.OK

    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def add_notification(user: str, title: str, message: str) -> Tuple[Dict[str, str], int]:
    """
    Adds a new notification for a user.

    :param user: The username.
    :param title: The title of the notification.
    :param message: The notification message.
    :return: A message and HTTP status code.
    """
    try:
        if 'notifications' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR

        new_notification = Notifications(user=user, title=title, notification=message)
        db.session.add(new_notification)
        db.session.commit()
        return {'message': 'Notification added'}, HTTPStatus.OK
    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_notifications(user_: str) -> Tuple[List[Notifications], int]:
    """
    Retrieves all notifications for a given user.

    :param user_: The username.
    :return: List of notifications and HTTP status code.
    """
    try:
        if 'notifications' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR
        
        notifications = Notifications.query.filter_by(user=user_).all()
        return notifications, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def delete_notification(user: str) -> Tuple[Dict[str, str], int]:
    """
    Deletes all notifications for a given user.

    :param user: The username.
    :return: A message and HTTP status code.
    """
    try:
        rows_deleted = Notifications.query.filter_by(user=user).delete()
        if rows_deleted == 0:
            return {'message': 'No notifications found for this user'}, HTTPStatus.NOT_FOUND

        db.session.commit()
        return {'message': 'deleted'}, HTTPStatus.OK

    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def add_job(data: dict) -> Tuple[Dict[str, str], int]:
    """
    Adds a new job to the database.

    :param data: Data containing details of the job.
    :return: A success message or error details.
    """
    try:
        contract_number = data.contract_number.data
        salesman = data.salesman.data
        customer = data.customer.data
        address = data.address.data
        city = data.city.data
        phone = data.phone.data
        description = data.description.data
        date = data.date.data
        email = data.email.data
        selling_price = data.selling_price.data
        cost = data.cost.data
        installer = data.installer.data
        status = data.status.data
        notes = data.notes.data

        new_job = Jobs(
            contract_number=contract_number,
            salesman=salesman,
            customer=customer,
            address=address,
            city=city,
            phone=phone,
            description=description,
            date=date,
            email=email,
            selling_price=selling_price,
            cost=cost,
            installer=installer,
            status=status,
            notes=notes
        )
        db.session.add(new_job)
        db.session.commit()
        return {'message': 'Job added'}, HTTPStatus.OK
    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def filter_job_salesman(data: dict) -> Tuple[List[Jobs], int]:
    """
    Filters jobs by salesman.

    :param data: Contains the salesman to filter by.
    :return: A list of jobs and HTTP status code.
    """
    try:
        jobs = Jobs.query.filter_by(salesman=data.salesman).all()
        return jobs, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_jobs(salesman: str, admin: bool) -> Tuple[List[Jobs], int]:
    """
    Retrieves jobs, either all or filtered by salesman.

    :param salesman: The salesman to filter by.
    :param admin: Flag to indicate if admin access.
    :return: A list of jobs and HTTP status code.
    """
    try:
        if 'jobs' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR
        if admin:
            jobs = Jobs.query.all()
        else:
            jobs = Jobs.query.filter_by(salesman=salesman).all()
        return jobs, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def get_the_latest_job() -> Tuple[List[Jobs], int]:
    """
    Retrieves the latest job.

    :return: The latest job.
    """
    try:
        if 'jobs' not in inspect(db.engine).get_table_names():
            return {'message': 'An error occurred', 'error': 'The table is not initialized. Please click the "Initialize" button on the main page.'}, HTTPStatus.INTERNAL_SERVER_ERROR

        job = Jobs.query.order_by(Jobs.id.desc()).first()
        return job, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def edit_job(data: dict, id_: int) -> Tuple[Dict[str, str], int]:
    """
    Edits a specific job based on ID.

    :param data: Contains the job details to edit.
    :param id_: ID of the job to be edited.
    :return: A success message or error details.
    """
    try:
        job = Jobs.query.get(id_)
        job.phone = data.phone.data
        job.description = data.description.data
        job.date = data.date.data
        job.email = data.email.data
        job.selling_price = data.selling_price.data
        job.cost = data.cost.data
        job.installer = data.installer.data
        job.status = data.status.data
        job.notes = data.notes.data
        db.session.commit()
        return {'message': 'Job edited'}, HTTPStatus.OK
    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def delete_job(id: int, send_exit_email: bool, email: str = None) -> Tuple[Dict[str, str], int]:
    """
    Deletes a job and optionally sends an exit email.

    :param id: The ID of the job to be deleted.
    :param send_exit_email: Flag indicating whether to send an exit email.
    :param email: The email address to send to.
    :return: A success message or error details.
    """
    conf = None
    if send_exit_email and email:
        builder = email_builder_factory('exit')
        email_sender = EmailSender(builder)
        email_sender.send_email(email, 'exit')
    job = Jobs.query.get(id)
    customer_ = Customer(
        contract_number=job.contract_number,
        salesman=job.salesman,
        name=job.customer,
        address=job.address,
        city=job.city,
        phone=job.phone,
        description=job.description,
        date=job.date,
        email=job.email
    )
    db.session.add(customer_)
    try:
        rows_deleted = Jobs.query.filter_by(id=id).delete()
        if rows_deleted == 0:
            return {'message': 'No job found with this id'}, HTTPStatus.NOT_FOUND

        db.session.commit()
        if send_exit_email and conf:
            return {'message': 'deleted, exit email sent'}, HTTPStatus.OK
        else:
            return {'message': 'deleted'}, HTTPStatus.OK
    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def send_quote_email(data, notes: str) -> Tuple[Dict[str, str], int]:
    """
    Sends a quote email and adds the quote to the database.

    :param data: Contains the quote details.
    :param notes: Additional notes for the quote.
    :return: A success message or error details.
    """
    quote = Quote(
        job_title=data.job_title.data,
        job_description=data.job_description.data,
        quote_amount=data.quote_amount.data,
        customer_name=data.customer_name.data,
        customer_address=data.customer_address.data,
        customer_email=data.customer_email.data,
        customer_phone=data.customer_phone.data,
        your_email=data.your_email.data,
        your_phone=data.your_phone.data,
        notes=data.notes.data
    )
    db.session.add(quote)
    db.session.commit()
    builder = email_builder_factory('quote')
    email_sender = EmailSender(builder)
    conf = email_sender.send_email(quote.customer_email, 'quote', data=quote, notes=notes)
    return {'message': 'Quote email sent', 'code': str(conf)}, HTTPStatus.OK

def get_quotes() -> Tuple[List[Quote], int]:
    """
    Retrieves all quotes from the database.

    :return: A list of quotes and HTTP status code.
    """
    try:
        quotes = Quote.query.all()
        return quotes, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def delete_quote(id: int) -> Tuple[Dict[str, str], int]:
    """
    Deletes a quote from the database based on the given ID.

    :param id: The ID of the quote to be deleted.
    :return: A success message or error details.
    """
    try:
        rows_deleted = Quote.query.filter_by(id=id).delete()
        if rows_deleted == 0:
            return {'message': 'No quote found with this id'}, HTTPStatus.NOT_FOUND

        db.session.commit()
        return {'message': 'deleted'}, HTTPStatus.OK

    except SQLAlchemyError as e:
        db.session.rollback()
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR

def resend_quote_email(id: int) -> Tuple[Dict[str, str], int]:
    """
    Resends a quote email based on the given quote ID.

    :param id: The ID of the quote whose email is to be resent.
    :return: A success message or error details.
    """
    try:
        quote = Quote.query.get(id)
        notes = quote.notes.split('<>')
        builder = email_builder_factory('quote')
        email_sender = EmailSender(builder)
        conf = email_sender.send_email(quote.customer_email, 'quote', data=quote, notes=notes)
        return {'message': 'Quote sent. Again!', 'email': str(conf)}, HTTPStatus.OK
    except SQLAlchemyError as e:
        return {'message': 'An error occurred', 'error': str(e)}, HTTPStatus.INTERNAL_SERVER_ERROR
    
def upload_spreadsheet_single(file: Any) -> None:
    """
    Uploads a single xls/xlsx file from the spreadsheet.gary page.

    :param file: The file to be uploaded, should have filename attribute ending with .xls or .xlsx
    :type file: Any
    :return: None
    """
    try:
        if file and (file.filename.endswith(".xls") or file.filename.endswith(".xlsx")):
            file.seek(0)
            stream = BytesIO(file.read())

            salesman = file.filename.split('.')[0]
            # read the file
            wb = load_workbook(stream, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                df = pd.DataFrame(sheet.values)
                header_row = df.iloc[0]
                
                # Iterate over the rows in DataFrame and add each as a new Customer
                for index, row in df.iterrows():
                    try:
                        if (row == header_row).all():
                            continue  
                        if row.isnull().all():
                            continue 

                        contract_number = str(row.iloc[0])[:64] if len(row) > 0 else None
                        name = str(row.iloc[1])[:64] if len(row) > 1 else None
                        address = str(row.iloc[2])[:120] if len(row) > 2 else None
                        city = str(row.iloc[3])[:64] if len(row) > 3 else None
                        phone = str(row.iloc[4])[:20] if len(row) > 4 else None
                        description = str(row.iloc[5])[:120] if len(row) > 5 else None
                        date = str(row.iloc[6])[:30] if len(row) > 6 else None
                        email = str(row.iloc[7])[:64] if len(row) > 7 and row.iloc[7] else 'NONE'

                        customer = Customer(
                            contract_number=contract_number,    
                            salesman=salesman,                       
                            name=name,               
                            address=address,           
                            city=city,               
                            phone=phone,              
                            description=description,       
                            date=date,                
                            email=email
                        )

                        db.session.add(customer)
                    except Exception as e:
                        print(e)
                        pass

            db.session.commit()
            print("A new sheet was added to the database")
    except Exception as e:
        print("An error occurred while processing the file: ", e)

def upload_spreadsheets_dump() -> None:
    """
    Dumps all xls/xlsx files in the app/spreadsheets folder.

    :return: None
    """
    directory = 'app/spreadsheets'
    for filename in os.listdir(directory):
        if file and (file.filename.endswith(".xls") or file.filename.endswith(".xlsx")):
            # File is an in-memory file object, need to convert it to an in-memory bytes stream object.
            file.seek(0)
            stream = BytesIO(file.read())

            salesman = file.filename.split('.')[0]
            # read the file
            wb = load_workbook(stream, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                df = pd.DataFrame(sheet.values)
                header_row = df.iloc[0]
                
                # Iterate over the rows in DataFrame and add each as a new Customer
                for index, row in df.iterrows():
                    try:
                        if (row == header_row).all():
                            continue  # Skip this row and move to the next one
                        if row.isnull().all():
                            continue  # Skip this row and move to the next one

                        contract_number = str(row.iloc[0])[:64] if len(row) > 0 else None
                        name = str(row.iloc[1])[:64] if len(row) > 1 else None
                        address = str(row.iloc[2])[:120] if len(row) > 2 else None
                        city = str(row.iloc[3])[:64] if len(row) > 3 else None
                        phone = str(row.iloc[4])[:20] if len(row) > 4 else None
                        description = str(row.iloc[5])[:120] if len(row) > 5 else None
                        date = str(row.iloc[6])[:30] if len(row) > 6 else None
                        email = str(row.iloc[7])[:64] if len(row) > 7 and row.iloc[7] else 'NONE'

                        customer = Customer(
                            contract_number=contract_number,    
                            salesman=salesman,                       
                            name=name,               
                            address=address,           
                            city=city,               
                            phone=phone,              
                            description=description,       
                            date=date,                
                            email=email
                        )

                        db.session.add(customer)
                    except Exception as e:
                        print(e)
                        pass
            db.session.commit()
        os.remove(file_path)

def check_if_siib_key_works(api_key: str) -> bool:
    """
    Checks if the SendinBlue API key works.

    :param api_key: The SendinBlue API key
    :return: True if the key works, False otherwise.
    """
    configuration = Configuration()
    configuration.api_key['api-key'] = api_key

    api_client = ApiClient(configuration)
    api_instance = AccountApi(api_client)

    try:
        # Fetches the account information to test the API key
        account_info = api_instance.get_account()
        return True
    except Exception:
        return False

def check_health() -> Dict[str, str]:
    """
    Checks the health of the database connection.

    :return: The status of the connection along with any related error codes.
    """
    db_status = 'connected'
    code = "NONE"
    try:
        with db.engine.connect() as connection:
            connection.execute(text('SELECT 1'))
    except Exception as e:
        db_status = 'disconnected'
        code = "DB ERROR CODE: " + str(e)
    return {
        'status': 'online',
        'database': db_status,
        'code': code
    }

def logout() -> None:
    """
    Logs out the current user.

    :return: None
    """
    logout_user()
